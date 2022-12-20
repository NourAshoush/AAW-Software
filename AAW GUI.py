from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook

root = Tk()
root.title("AAW Monthly Sales Editor")
root.geometry("500x500")
root.resizable(0, 0)
root.iconbitmap("Logo-Bioderma.ico")

uploadOriginFileStr = ""
uploadDestFileStr = ""
readFromColumn = ""
writeToColumn = ""


def uploadOriginFile():
    global uploadOriginFileStr
    uploadOriginFileStr = str(filedialog.askopenfilename())
    if uploadOriginFileStr != "":
        uploadOriginButton.config(text=(uploadOriginFileStr.split("/"))
                                  [len(uploadOriginFileStr.split("/")) - 1])
        enable(readFromLabel)
        enable(readFromText)
        checkState()
    else:
        uploadOriginButton.config(text="Select File")
        disable(readFromLabel)
        readFromText.delete(0, END)
        disable(readFromText)


def uploadDestFile():
    global uploadDestFileStr
    uploadDestFileStr = str(filedialog.askopenfilename())
    if uploadDestFileStr != "":
        uploadDestButton.config(text=(uploadDestFileStr.split("/"))
                                [len(uploadDestFileStr.split("/")) - 1])
        enable(writeToLabel)
        enable(writeToText)
        checkState()
    else:
        uploadDestButton.config(text="Select File")
        disable(writeToLabel)
        writeToText.delete(0, END)
        disable(writeToText)


def start():
    if checkText():
        statusLabel.config(text="Running...", fg="green")
        disable(uploadOriginButton)
        disable(uploadDestButton)
        disable(readFromText)
        disable(writeToText)
        disable(runButton)
        root.update()
        try:
            runCopy()
        except:
            statusLabel.config(text="An error has occurred!", fg="red")
        else:
            statusLabel.config(text="Done!", font=(
                "Arial bold", 15), fg="green")
        finally:
            enable(uploadOriginButton)
            enable(uploadDestButton)
            enable(readFromText)
            enable(writeToText)
            enable(runButton)


def disable(widget):
    widget.config(state=DISABLED)


def enable(widget):
    widget.config(state=NORMAL)


def checkState():
    if uploadOriginFileStr != "" and uploadDestFileStr != "":
        runButton.config(state=NORMAL)
    else:
        runButton.config(state=DISABLED)


def checkText():
    if readFromText.get().isalpha() and writeToText.get().isalpha():
        global readFromColumn, writeToColumn
        readFromColumn = (readFromText.get()).upper()
        writeToColumn = (writeToText.get()).upper()
        statusLabel.config(text="", fg="green")
        return True

    else:
        if not readFromText.get().isalpha() and not writeToText.get().isalpha():
            readFromText.delete(0, END)
            writeToText.delete(0, END)
            statusLabel.config(text="Please check both input columns")

        else:
            if not readFromText.get().isalpha():
                readFromText.delete(0, END)
                statusLabel.config(text="Please check `Sales Value` column")

            if not writeToText.get().isalpha():
                writeToText.delete(0, END)
                statusLabel.config(text="Please check `Month-Year` column")

        return False


def runCopy():
    book = load_workbook(uploadDestFileStr)
    sheet = book["VALUE"]

    rowIndex = 4
    verifyRowNum = "A4"
    codeAddress = "B4"
    codeList = []
    codeAddressList = []

    while str(sheet[verifyRowNum].value).isnumeric():

        if str(sheet[codeAddress].value).isnumeric():
            codeList.append(str(sheet[codeAddress].value))
            codeAddressList.append(
                writeToColumn + str(''.join([i for i in codeAddress if i.isdigit()])))

        rowIndex += 1
        verifyRowNum = "A" + str(rowIndex)
        codeAddress = "B" + str(rowIndex)

    book = load_workbook(uploadOriginFileStr)
    sheet = book["By Customer"]

    originMaxRow = sheet.max_row
    bColumn = []
    valueList = []

    for i in range(6, originMaxRow + 1):
        bColumn.append(sheet["B" + str(i)].value)

    for code in codeList:
        if code in bColumn:
            i = bColumn.index(code) + 1
            while bColumn[i] == None:
                i += 1
            i += 6
            valueList.append(sheet[readFromColumn + str(i)].value)
        else:
            valueList.append(None)

    book = load_workbook(uploadDestFileStr)
    sheet = book["VALUE"]

    for value, address in zip(valueList, codeAddressList):
        if value is not None:
            sheet[address] = value

    book.save(uploadDestFileStr)


title = Label(
    root, text="Copy from `By Customer` sheet to \n`Value` sheet", font=("Arial", 25))
title.place(x=0, y=30, width=500)

instructionsLabel = Label(root,
                          text="`Origin File` is the Excel file with the data to be copied.\n`Destination File` is the Excel file that will be edited.\nFor column inputs, enter only the column's letter(s). For example: 'A' , 'CD'.\nBefore running, please make sure that the `Origin File` and `Destination File` \nare closed.",
                          font=("Arial", 13),
                          fg="grey")
instructionsLabel.place(x=0, y=95, width=500)

originLabel = Label(root, text="Origin File", font=("Arial", 17))
originLabel.place(x=50, y=200, width=200)

destLabel = Label(root, text="Destination File", font=("Arial", 17))
destLabel.place(x=250, y=200, width=200)

uploadOriginButton = Button(root, text="Select File", command=uploadOriginFile)
uploadOriginButton.place(x=50, y=240, width=200)

uploadDestButton = Button(root, text="Select File", command=uploadDestFile)
uploadDestButton.place(x=250, y=240, width=200)

readFromLabel = Label(root, text="`Sales Value` Column",
                      state=DISABLED, font=("Arial", 17))
readFromLabel.place(x=50, y=300, width=200)

writeToLabel = Label(root, text="`Month-Year` Column",
                     state=DISABLED, font=("Arial", 17))
writeToLabel.place(x=250, y=300, width=200)

readFromText = Entry(root, state=DISABLED, justify=CENTER, border=2)
readFromText.place(x=100, y=340, width=100)

writeToText = Entry(root, state=DISABLED, justify=CENTER, border=2)
writeToText.place(x=300, y=340, width=100)

statusLabel = Label(root, text="", font=("Arial", 15), fg="red")
statusLabel.place(x=50, y=400, width=400)

runButton = Button(root, text="Run", command=start, state=DISABLED)
runButton.place(x=200, y=450, width=100, height=30)

root.bind('<Return>', lambda event=None: runButton.invoke())

root.bind('<Escape>', lambda event=None: root.destroy())

root.mainloop()
